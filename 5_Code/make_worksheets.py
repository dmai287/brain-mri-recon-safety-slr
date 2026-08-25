# -*- coding: utf-8 -*-
"""M: independent screening worksheets for the three human reviewers.
One identical workbook per reviewer covering ALL 264 included studies, the 44
newest additions first (they need first-time human confirmation; the 220 earlier
includes are re-verification). No formulas - pure data-entry forms."""
import csv
import pathlib
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

sys.stdout.reconfigure(encoding="utf-8", errors="replace")
BASE = pathlib.Path(r"C:\Users\V133280\RMIT University\Thai Pham - StrokeVault"
                    r"\Brain MRI Reconstruction Safety Review - Broad")
OUT = BASE / "Full Text Screening" / "Reviewer_Worksheets"
OUT.mkdir(exist_ok=True)

rows = list(csv.DictReader((BASE / "Registration/OSF_Upload/3_Included_Studies/"
                            "included_characteristics.csv").open(encoding="utf-8-sig")))
sup = {r["study_key"]: r for r in csv.DictReader(
    (BASE / "Registration/OSF_Upload/3_Included_Studies/"
     "included_characteristics_supplement.csv").open(encoding="utf-8-sig"))}
new44 = [r for r in rows if int(r["study_key"][1:]) > 220]
old220 = [r for r in rows if int(r["study_key"][1:]) <= 220]
ordered = new44 + old220

REVIEWERS = ["Dat_Tat_Mai", "Thai_Viet_Pham", "Thu_Nguyen_Thi_Dang"]
ARIAL = Font(name="Arial", size=10)
BOLD = Font(name="Arial", size=10, bold=True)
HDR_FILL = PatternFill("solid", fgColor="D9E2E2")
INPUT_FILL = PatternFill("solid", fgColor="FFFF99")
PRIO_FILL = PatternFill("solid", fgColor="E6F2F1")
EX_FILL = PatternFill("solid", fgColor="EEEEEE")
WRAP = Alignment(wrap_text=True, vertical="top")

HEADERS = ["#", "study_key", "priority", "year", "title", "journal", "doi_link",
           "full_text_note",
           "P: human brain MRI? (Y/N/U)", "I: recon/accel method or its evaluation? (Y/N/U)",
           "O: fidelity/failure/observer outcome? (Y/N/U)", "D: primary empirical study? (Y/N/U)",
           "DECISION (include/exclude/unsure)", "exclusion reason (if exclude)",
           "supporting quote (verbatim)", "notes"]
WIDTHS = [5, 10, 16, 6, 52, 30, 34, 34, 14, 14, 14, 14, 16, 26, 40, 26]
DEC = DataValidation(type="list", formula1='"include,exclude,unsure"', allow_blank=True)
YNU = DataValidation(type="list", formula1='"Y,N,U"', allow_blank=True)
REASONS = ('"out_of_scope,wrong_intervention_or_index,wrong_outcome,'
           'wrong_population,wrong_study_design,non_english,'
           'review_editorial_protocol_commentary"')
REA = DataValidation(type="list", formula1=REASONS, allow_blank=True)

for rev in REVIEWERS:
    wb = Workbook()
    ws = wb.active
    ws.title = "Instructions"
    lines = [
        ("Independent full-text screening worksheet", True),
        (f"Reviewer: {rev.replace('_', ' ')}", True),
        ("", False),
        ("Scope: all 264 included studies of the review. The first 44 rows "
         "(priority = CONFIRM-FIRST, tinted) are the studies added by the final "
         "retrieval round and need your first-time eligibility decision; the "
         "remaining 220 are re-verification of earlier decisions.", False),
        ("", False),
        ("How to complete:", True),
        ("1. Work independently. Do not consult the other reviewers or any prior "
         "decision file until all three worksheets are returned.", False),
        ("2. Open the study's full text (the full_text_note column names the note "
         "file in Full Text Screening/Include/; the doi_link opens the publisher "
         "page).", False),
        ("3. Answer the four PICOS columns Y/N/U, then set DECISION. Only the "
         "YELLOW columns are yours to edit.", False),
        ("4. If exclude: pick an exclusion reason and paste a short verbatim quote "
         "that supports it. If include: a supporting quote is encouraged for the "
         "44 priority rows.", False),
        ("5. Return the completed file. Decisions are integrated as the human "
         "screening record; disagreements are resolved by discussion, and "
         "inter-rater agreement will be computed from the three independent "
         "sheets.", False),
        ("", False),
        ("Eligibility summary (full criteria: PICOS table of the manuscript):", True),
        ("P - human brain MRI acquisitions. I - deep-learning/accelerated "
         "reconstruction methods, or methods for evaluating their fidelity, "
         "artifacts, robustness or failure modes. O - image-quality/fidelity, "
         "artifact or failure-mode characterisation, metric-reader agreement, "
         "uncertainty, robustness, or diagnostic/observer outcomes. D - primary "
         "empirical study (no reviews/editorials/protocols).", False),
        ("", False),
        ("Legend: YELLOW = enter your answer. Grey EXAMPLE row shows the expected "
         "format - do not edit it. Tinted rows = the 44 priority studies.", True),
    ]
    for i, (txt, bold) in enumerate(lines, 1):
        c = ws.cell(row=i, column=1, value=txt)
        c.font = BOLD if bold else ARIAL
        c.alignment = WRAP
    ws.column_dimensions["A"].width = 110

    ws2 = wb.create_sheet("Screening")
    for j, (h, w) in enumerate(zip(HEADERS, WIDTHS), 1):
        c = ws2.cell(row=1, column=j, value=h)
        c.font = BOLD
        c.fill = HDR_FILL
        c.alignment = WRAP
        ws2.column_dimensions[get_column_letter(j)].width = w
    ws2.add_data_validation(DEC)
    ws2.add_data_validation(YNU)
    ws2.add_data_validation(REA)

    # example row (grey, marked, not a real study)
    ex = ["EX", "EXAMPLE", "-", "2024", "EXAMPLE ROW - expected format, do not edit",
          "Example Journal", "https://doi.org/10.xxxx/example",
          "2024 - Example study - PMID 0.md",
          "Y", "Y", "Y", "Y", "include", "",
          "\"readers rated diagnostic quality of the accelerated scans\"",
          "example only"]
    for j, v in enumerate(ex, 1):
        c = ws2.cell(row=2, column=j, value=v)
        c.font = ARIAL
        c.fill = EX_FILL
        c.alignment = WRAP

    r_i = 3
    for n, r in enumerate(ordered, 1):
        prio = "CONFIRM-FIRST" if int(r["study_key"][1:]) > 220 else "re-verify"
        reader_pending = (sup.get(r["study_key"], {}).get("rater_status", "")
                          .startswith("pending"))
        note = r["source_record"]
        vals = [n, r["study_key"], prio, r["year"], r["title"], r["journal"],
                ("https://doi.org/" + r["doi"]) if r["doi"] else "",
                note, "", "", "", "", "", "", "",
                ("reader-design: QUADAS-2 appraisal also pending" if reader_pending
                 else "")]
        for j, v in enumerate(vals, 1):
            c = ws2.cell(row=r_i, column=j, value=v)
            c.font = ARIAL
            c.alignment = WRAP
            if j in (9, 10, 11, 12, 13, 14, 15):
                c.fill = INPUT_FILL
            elif prio == "CONFIRM-FIRST":
                c.fill = PRIO_FILL
        DEC.add(ws2.cell(row=r_i, column=13))
        REA.add(ws2.cell(row=r_i, column=14))
        for col in (9, 10, 11, 12):
            YNU.add(ws2.cell(row=r_i, column=col))
        r_i += 1
    ws2.freeze_panes = "A2"

    f = OUT / f"screening_worksheet_264_{rev}.xlsx"
    wb.save(str(f))
    print("wrote", f.name, f"({r_i - 3} study rows)")

print("done ->", OUT)
