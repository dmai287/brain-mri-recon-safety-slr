# -*- coding: utf-8 -*-
"""Q: QUADAS-2 appraisal forms for the 18 pending reader-design studies.
One workbook per reviewer (independent), long format: one row per study x domain
(18 x 7 = 126 rows), with the review's tailored signalling questions as guidance,
Low/Unclear/High dropdowns, and mandatory verbatim-evidence + rationale cells -
the same evidence-quotation rule as the 39 appraised studies."""
import csv
import pathlib
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

sys.stdout.reconfigure(encoding="utf-8", errors="replace")
BASE = pathlib.Path(r"C:\Users\V133280\RMIT University\Thai Pham - StrokeVault"
                    r"\Brain MRI Reconstruction Safety Review - Broad")
OUT = BASE / "Full Text Screening" / "Reviewer_Worksheets"

chars = {r["study_key"]: r for r in csv.DictReader(
    (BASE / "Registration/OSF_Upload/3_Included_Studies/"
     "included_characteristics.csv").open(encoding="utf-8-sig"))}
sup = [r for r in csv.DictReader(
    (BASE / "Registration/OSF_Upload/3_Included_Studies/"
     "included_characteristics_supplement.csv").open(encoding="utf-8-sig"))
    if r["rater_status"].startswith("pending")]
assert len(sup) == 18

DOMAINS = [
    ("rob_patient_selection", "Risk of bias - D1 Patient Selection",
     "Was a consecutive or random sample of patients/scans enrolled? Was a "
     "case-control design avoided? Were inappropriate exclusions avoided? For "
     "reader studies of reconstruction: is the rule for selecting the scans that "
     "readers saw stated (consecutive, random, purposive, pathology-enriched)?"),
    ("rob_index_test", "Risk of bias - D2 Index Test",
     "Were the index-test results interpreted without knowledge of the reference "
     "standard? For reconstruction reader studies: were readers blinded to the "
     "reconstruction condition (method/acceleration), was presentation order "
     "randomised, and were rating thresholds pre-specified? Note High if the "
     "evaluated method is post-processing/synthesis rather than reconstruction "
     "from undersampled k-space (per the review's rule)."),
    ("rob_reference_standard", "Risk of bias - D3 Reference Standard",
     "Is the reference standard likely to correctly classify / represent ground "
     "truth (fully sampled acquisition, conventional-of-care sequence, expert "
     "consensus)? Was it interpreted without knowledge of the index test? Is a "
     "motion/artifact-free reference assumption tenable here?"),
    ("rob_flow_timing", "Risk of bias - D4 Flow and Timing",
     "Did all scans receive the same reference standard? Were all enrolled "
     "scans/patients included in the analysis (dropouts, unreadable exams "
     "accounted for)? Appropriate interval / same-session comparison?"),
    ("app_patient_selection", "Applicability - A1 Patient Selection",
     "Do the included patients/scans match the review question: clinical human "
     "brain MRI (not phantom-only, not non-brain)?"),
    ("app_index_test", "Applicability - A2 Index Test",
     "Does the evaluated reconstruction/acceleration and its reader evaluation "
     "match the review question (deep-learning or accelerated brain MRI "
     "reconstruction and its fidelity/diagnostic evaluation)?"),
    ("app_reference_standard", "Applicability - A3 Reference Standard",
     "Does the reference standard match how the target condition/image quality "
     "would be established in practice?"),
]

REVIEWERS = ["Dat_Tat_Mai", "Thai_Viet_Pham", "Thu_Nguyen_Thi_Dang"]
ARIAL = Font(name="Arial", size=10)
BOLD = Font(name="Arial", size=10, bold=True)
HDR = PatternFill("solid", fgColor="D9E2E2")
YEL = PatternFill("solid", fgColor="FFFF99")
GREY = PatternFill("solid", fgColor="EEEEEE")
BAND = PatternFill("solid", fgColor="E6F2F1")
WRAP = Alignment(wrap_text=True, vertical="top")

for rev in REVIEWERS:
    wb = Workbook()
    ws = wb.active
    ws.title = "Instructions"
    lines = [
        ("QUADAS-2 appraisal - 18 reader-design studies pending appraisal", True),
        (f"Reviewer: {rev.replace('_', ' ')}", True),
        ("", False),
        ("These 18 studies carry rater_status = pending in "
         "included_characteristics_supplement.csv. Appraise them with the same "
         "instrument and rules as the 39 already-appraised studies:", False),
        ("1. Work independently; do not consult the other reviewers until all "
         "three workbooks are returned.", False),
        ("2. Rate every domain Low / Unclear / High concern.", False),
        ("3. Every rating MUST carry a verbatim supporting quotation from the "
         "study's full text (the note file named in the Studies sheet) and a "
         "one-sentence rationale. No quote, no rating.", False),
        ("4. If a domain cannot be judged from the text, rate Unclear and quote "
         "the closest passage (or state 'not reported').", False),
        ("5. If full-text appraisal finds NO human reader assessment, write "
         "NO-READER in the rationale of D2 and stop; the study will be "
         "reclassified as algorithmic (this happened to 4 studies before).", False),
        ("6. Return the completed file; ratings are integrated only after all "
         "three agree or disagreements are resolved by discussion, exactly as "
         "for the earlier batches.", False),
        ("", False),
        ("Only YELLOW cells are yours to edit. The Studies sheet lists each "
         "study's full-text note, DOI and extracted reader count for reference.",
         True),
    ]
    for i, (txt, bold) in enumerate(lines, 1):
        c = ws.cell(row=i, column=1, value=txt)
        c.font = BOLD if bold else ARIAL
        c.alignment = WRAP
    ws.column_dimensions["A"].width = 110

    # Studies reference sheet
    ws1 = wb.create_sheet("Studies")
    heads = ["study_key", "year", "title", "journal", "doi_link",
             "full_text_note (Full Text Screening/Include/)",
             "n_readers (extracted)", "design_detail"]
    widths = [10, 6, 56, 26, 34, 44, 12, 26]
    for j, (h, w) in enumerate(zip(heads, widths), 1):
        c = ws1.cell(row=1, column=j, value=h)
        c.font = BOLD
        c.fill = HDR
        c.alignment = WRAP
        ws1.column_dimensions[get_column_letter(j)].width = w
    for i, r in enumerate(sup, 2):
        ch = chars[r["study_key"]]
        vals = [r["study_key"], r["year"], r["title"], ch["journal"],
                ("https://doi.org/" + r["doi"]) if r["doi"] else "",
                ch["source_record"], r["n_readers"], r["design_detail"]]
        for j, v in enumerate(vals, 1):
            c = ws1.cell(row=i, column=j, value=v)
            c.font = ARIAL
            c.alignment = WRAP
    ws1.freeze_panes = "A2"

    # Appraisal sheet: long format
    ws2 = wb.create_sheet("Appraisal")
    heads = ["study_key", "title (short)", "domain", "signalling questions (guidance)",
             "RATING (Low/Unclear/High)", "verbatim supporting quotation",
             "rationale (one sentence)"]
    widths = [10, 34, 30, 56, 16, 46, 34]
    for j, (h, w) in enumerate(zip(heads, widths), 1):
        c = ws2.cell(row=1, column=j, value=h)
        c.font = BOLD
        c.fill = HDR
        c.alignment = WRAP
        ws2.column_dimensions[get_column_letter(j)].width = w
    RATE = DataValidation(type="list", formula1='"Low,Unclear,High"',
                          allow_blank=True)
    ws2.add_data_validation(RATE)
    ri = 2
    for si, r in enumerate(sup):
        band = si % 2 == 0
        for key, dom, guide in DOMAINS:
            vals = [r["study_key"], r["title"][:60], dom, guide, "", "", ""]
            for j, v in enumerate(vals, 1):
                c = ws2.cell(row=ri, column=j, value=v)
                c.font = ARIAL
                c.alignment = WRAP
                if j in (5, 6, 7):
                    c.fill = YEL
                elif band:
                    c.fill = BAND
            RATE.add(ws2.cell(row=ri, column=5))
            ri += 1
    ws2.freeze_panes = "D2"

    f = OUT / f"quadas2_worksheet_18_{rev}.xlsx"
    wb.save(str(f))
    print("wrote", f.name, f"({ri - 2} domain rows)")
print("done ->", OUT)
