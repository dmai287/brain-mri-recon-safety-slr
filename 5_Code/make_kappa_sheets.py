# -*- coding: utf-8 -*-
"""N: blinded kappa worksheets (reviewer M4). Random 20% of the 1,068
full-text-assessed reports (n=214), stratified on the recorded decision so the
sample mirrors the corpus mix, then SHUFFLED and stripped of any prior-decision
hint. Fixed seed (2026) so the draw is reproducible and documented. The key
(row -> recorded decision) is written to a separate file the reviewers must not
open until all three sheets are returned."""
import csv
import pathlib
import random
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

sys.stdout.reconfigure(encoding="utf-8", errors="replace")
BASE = pathlib.Path(r"C:\Users\V133280\RMIT University\Thai Pham - StrokeVault"
                    r"\Brain MRI Reconstruction Safety Review - Broad")
OUT = BASE / "Full Text Screening" / "Reviewer_Worksheets"
OUT.mkdir(exist_ok=True)

rows = list(csv.DictReader((BASE / "Full Text Screening/"
                            "agent3_final_screening_decisions.csv").open(encoding="utf-8-sig")))
assert len(rows) == 1068
rng = random.Random(2026)
TARGET = 214  # 20% of 1,068 (M4's spec)

by_dec = {}
for r in rows:
    by_dec.setdefault(r["decision"], []).append(r)
sample = []
for dec, pool in by_dec.items():
    k = round(TARGET * len(pool) / len(rows))
    sample += rng.sample(pool, k)
while len(sample) < TARGET:
    extra = rng.choice(rows)
    if extra not in sample:
        sample.append(extra)
sample = sample[:TARGET]
rng.shuffle(sample)
from collections import Counter
mix = Counter(r["decision"] for r in sample)
print(f"sample n={len(sample)} | mix (kept secret from reviewers): {dict(mix)}")

# key file (sealed)
with (OUT / "_KAPPA_KEY_do_not_open_until_all_sheets_returned.csv"
      ).open("w", newline="", encoding="utf-8-sig") as fh:
    w = csv.writer(fh)
    w.writerow(["row_id", "doi", "pmid", "title", "recorded_decision", "stage"])
    for i, r in enumerate(sample, 1):
        w.writerow([f"K{i:03d}", r["doi"], r["pmid"], r["title"],
                    r["decision"], r["stage"]])
print("key written (sealed file)")

REVIEWERS = ["Dat_Tat_Mai", "Thai_Viet_Pham", "Thu_Nguyen_Thi_Dang"]
ARIAL = Font(name="Arial", size=10)
BOLD = Font(name="Arial", size=10, bold=True)
HDR_FILL = PatternFill("solid", fgColor="D9E2E2")
INPUT_FILL = PatternFill("solid", fgColor="FFFF99")
EX_FILL = PatternFill("solid", fgColor="EEEEEE")
WRAP = Alignment(wrap_text=True, vertical="top")
HEADERS = ["row_id", "year", "title", "journal", "doi_link",
           "P: human brain MRI? (Y/N/U)",
           "I: recon/accel method or its evaluation? (Y/N/U)",
           "O: fidelity/failure/observer outcome? (Y/N/U)",
           "D: primary empirical study? (Y/N/U)",
           "DECISION (include/exclude/unsure)",
           "exclusion reason (if exclude)", "supporting quote (verbatim)", "notes"]
WIDTHS = [8, 6, 56, 30, 34, 14, 14, 14, 14, 16, 26, 40, 24]
REASONS = ('"out_of_scope,wrong_intervention_or_index,wrong_outcome,'
           'wrong_population,wrong_study_design,non_english,'
           'review_editorial_protocol_commentary"')

for rev in REVIEWERS:
    wb = Workbook()
    ws = wb.active
    ws.title = "Instructions"
    lines = [
        ("Blinded eligibility re-screening for inter-rater agreement (kappa)", True),
        (f"Reviewer: {rev.replace('_', ' ')}", True),
        ("", False),
        ("This sheet holds a random 20% sample (n=214, fixed seed 2026) of ALL "
         "reports assessed at full text - a mix of studies that were included and "
         "reports that were excluded, in random order. You are NOT told which is "
         "which, and the mix proportions are withheld.", False),
        ("", False),
        ("Rules:", True),
        ("1. Decide every row afresh against the PICOS criteria. Do not look up "
         "prior decision files, do not consult the other reviewers, and do not "
         "open the sealed key file, until all three sheets are returned.", False),
        ("2. Only the YELLOW columns are yours to edit.", False),
        ("3. When all three sheets are back, Fleiss' kappa across the three raters "
         "and each rater's agreement with the recorded decisions will be computed "
         "and reported exactly as measured.", False),
        ("", False),
        ("Eligibility summary: P - human brain MRI acquisitions. I - deep-learning/"
         "accelerated reconstruction methods, or methods for evaluating their "
         "fidelity, artifacts, robustness or failure modes. O - image-quality/"
         "fidelity, artifact or failure-mode characterisation, metric-reader "
         "agreement, uncertainty, robustness, or diagnostic/observer outcomes. "
         "D - primary empirical study (no reviews/editorials/protocols).", False),
        ("", False),
        ("Legend: YELLOW = enter your answer. Grey EXAMPLE row shows the format - "
         "do not edit it.", True),
    ]
    for i, (txt, bold) in enumerate(lines, 1):
        c = ws.cell(row=i, column=1, value=txt)
        c.font = BOLD if bold else ARIAL
        c.alignment = WRAP
    ws.column_dimensions["A"].width = 110

    ws2 = wb.create_sheet("Kappa_sample")
    DEC = DataValidation(type="list", formula1='"include,exclude,unsure"',
                         allow_blank=True)
    YNU = DataValidation(type="list", formula1='"Y,N,U"', allow_blank=True)
    REA = DataValidation(type="list", formula1=REASONS, allow_blank=True)
    ws2.add_data_validation(DEC)
    ws2.add_data_validation(YNU)
    ws2.add_data_validation(REA)
    for j, (h, w) in enumerate(zip(HEADERS, WIDTHS), 1):
        c = ws2.cell(row=1, column=j, value=h)
        c.font = BOLD
        c.fill = HDR_FILL
        c.alignment = WRAP
        ws2.column_dimensions[get_column_letter(j)].width = w
    ex = ["EX", "2024", "EXAMPLE ROW - expected format, do not edit",
          "Example Journal", "https://doi.org/10.xxxx/example",
          "Y", "Y", "N", "Y", "exclude", "wrong_outcome",
          "\"reports segmentation accuracy only; no reconstruction outcome\"",
          "example only"]
    for j, v in enumerate(ex, 1):
        c = ws2.cell(row=2, column=j, value=v)
        c.font = ARIAL
        c.fill = EX_FILL
        c.alignment = WRAP
    r_i = 3
    for i, r in enumerate(sample, 1):
        year = (r.get("year") or "").split(".")[0]
        vals = [f"K{i:03d}", year, r["title"], r["journal"],
                ("https://doi.org/" + r["doi"]) if r["doi"] else
                (f"https://pubmed.ncbi.nlm.nih.gov/{r['pmid']}/" if r["pmid"] else ""),
                "", "", "", "", "", "", "", ""]
        for j, v in enumerate(vals, 1):
            c = ws2.cell(row=r_i, column=j, value=v)
            c.font = ARIAL
            c.alignment = WRAP
            if j >= 6:
                c.fill = INPUT_FILL
        DEC.add(ws2.cell(row=r_i, column=10))
        REA.add(ws2.cell(row=r_i, column=11))
        for col in (6, 7, 8, 9):
            YNU.add(ws2.cell(row=r_i, column=col))
        r_i += 1
    ws2.freeze_panes = "A2"
    f = OUT / f"kappa_worksheet_214_{rev}.xlsx"
    wb.save(str(f))
    print("wrote", f.name)
print("done")
